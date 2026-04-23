#!/usr/bin/env python3
"""
paychex_to_401k.py — Generate Creative Planning 401(k) contribution files
from Paychex Retirement Plan Summary PDF.

Outputs to payroll folder:
  {prefix}-{YYYYMMDD}-001-CREATIVEPLANNING-DIRECTLIGHTING-GUID-{GUID}.csv
  {prefix}-{YYYYMMDD}-001-CREATIVEPLANNING-DIRECTLIGHTING-GUID-{GUID}.xlsx

Usage:
  python3 paychex_to_401k.py                             # auto-detects latest payroll folder
  python3 paychex_to_401k.py /path/to/Payroll.MMDD       # specific folder
  python3 paychex_to_401k.py --build-ref /path/to/template.xlsx  # rebuild employee roster
"""

import re
import os
import sys
import json
import datetime
from pathlib import Path

import pdfplumber
import openpyxl

SKILL_DIR = Path(__file__).parent
NAS_BASE = Path("/Volumes/Public/Direct Lighting/Direct Lighting LLC/1.Direct.Payroll/1.Payrolls")
REF_FILE = SKILL_DIR / "401k_employee_ref.json"

# Creative Planning fixed identifiers — update if CP issues a new GUID
META_PREFIX = "015-20141180-20220006-20220006"
CP_GUID = "00000001F2F3"


# ---------------------------------------------------------------------------
# Reference file management
# ---------------------------------------------------------------------------

def load_ref():
    if REF_FILE.exists():
        return json.loads(REF_FILE.read_text())
    return {"employees": []}


def save_ref(ref):
    REF_FILE.write_text(json.dumps(ref, indent=2))


def build_ref_from_xlsx(xlsx_path):
    """
    Build employee roster from a Creative Planning xlsx template.
    Keyed by SSN last-4. Preserves row order for consistent file output.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    employees = []
    for row in ws.iter_rows(values_only=True):
        ssn = str(row[0] or "").strip()
        if not re.match(r"\d{3}-\d{2}-\d{4}", ssn):
            continue
        employees.append({
            "last4_ssn": ssn[-4:],
            "ssn":       ssn,
            "first":     str(row[1]  or "").strip(),
            "mi":        str(row[2]  or "").strip(),
            "last":      str(row[3]  or "").strip(),
            "sex":       str(row[4]  or "").strip(),
            "dob":       str(row[5]  or "").strip(),
            "hire":      str(row[6]  or "").strip(),
            "addr1":     str(row[14] or "").strip(),
            "addr2":     str(row[15] or "").strip(),
            "city":      str(row[16] or "").strip(),
            "state":     str(row[17] or "").strip(),
            "zip":       str(row[18] or "").strip(),
        })
    return {"employees": employees}


# ---------------------------------------------------------------------------
# Payroll folder detection
# ---------------------------------------------------------------------------

def find_latest_payroll_folder():
    year = datetime.date.today().strftime("%y")  # e.g. '26'
    quarters = sorted(NAS_BASE.glob(f"{year}.Payrolls/Q*"))
    for q in reversed(quarters):
        folders = sorted(q.glob("Payroll.*"))
        if folders:
            return folders[-1]
    return None


def find_retirement_pdf(folder):
    pdfs = sorted(Path(folder).glob("*_Retirement_Plan_Summary_*.pdf"))
    return pdfs[0] if pdfs else None


# ---------------------------------------------------------------------------
# PDF parsing
# ---------------------------------------------------------------------------

def _parse_num(s):
    try:
        return float(s.replace(",", ""))
    except (ValueError, AttributeError):
        return 0.0


def parse_retirement_pdf(pdf_path):
    """
    Parse a Paychex Retirement Plan Summary PDF.

    Returns:
        check_date  — datetime.date from "CheckDate MM/DD/YY" in PDF
        contribs    — dict keyed by SSN last-4:
                      {hours, earnings, pretax, roth, employer}
    """
    with pdfplumber.open(pdf_path) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    # --- Check date ---
    check_date = None
    m = re.search(r"CheckDate\s+(\d{2}/\d{2}/\d{2})", full_text)
    if m:
        check_date = datetime.datetime.strptime(m.group(1), "%m/%d/%y").date()

    # --- Per-participant contributions ---
    contribs = {}
    lines = full_text.split("\n")
    in_participant_section = True
    current = None  # accumulates data for the employee being parsed

    for line in lines:
        line = line.strip()

        # Section boundary — stop when we hit totals or non-participant table
        if re.match(r"401\(k\)TOTALS|NON-PARTICIPANTINFORMATION", line):
            in_participant_section = False
            current = None
            continue

        if not in_participant_section:
            continue

        # --- New participant: "Lastname,FirstnameM ThisPeriod HOURS EARNINGS ..." ---
        tp = re.match(
            r"([A-Za-z'\-]+,[A-Za-z]+)\s+ThisPeriod\s+([\d\s,\.]+)",
            line,
        )
        if tp:
            name_raw, amounts_raw = tp.group(1), tp.group(2).strip()
            amounts = [_parse_num(x) for x in re.findall(r"[\d,]+\.\d{2}", amounts_raw)]
            current = {
                "last":    name_raw.split(",")[0],
                "amounts": amounts,
                "last4":   None,
                "type":    None,   # 'pretax' | 'roth' | 'both'
            }
            continue

        if current is None:
            continue

        # --- SSN (masked) ---
        ssn_m = re.search(r"SSN:\s*xxx-xx-(\d{4})", line)
        if ssn_m:
            current["last4"] = ssn_m.group(1)
            continue

        # --- Contribution type → finalise participant ---
        type_m = re.search(r"EmployeeRecurringAmount\(s\):(.*)", line)
        if type_m and current["last4"]:
            type_str = type_m.group(1)
            has_pretax = "Pre-tax" in type_str
            has_roth   = "Post-tax" in type_str or "CatchUp" in type_str

            amounts  = current["amounts"]
            hours    = amounts[0] if len(amounts) > 0 else 0.0
            earnings = amounts[1] if len(amounts) > 1 else 0.0
            contribs_raw = amounts[2:]

            pretax = roth = employer = 0.0
            if has_pretax and has_roth and len(contribs_raw) >= 3:
                pretax, roth, employer = contribs_raw[0], contribs_raw[1], contribs_raw[2]
            elif has_pretax and len(contribs_raw) >= 2:
                pretax, employer = contribs_raw[0], contribs_raw[1]
            elif has_roth and len(contribs_raw) >= 2:
                roth, employer = contribs_raw[0], contribs_raw[1]
            elif len(contribs_raw) == 1:
                employer = contribs_raw[0]

            contribs[current["last4"]] = {
                "hours":    hours,
                "earnings": earnings,
                "pretax":   pretax,
                "roth":     roth,
                "employer": employer,
            }
            current = None

    return check_date, contribs


# ---------------------------------------------------------------------------
# Output generation
# ---------------------------------------------------------------------------

def _clean(v):
    """Return int for whole floats, else leave as-is. None → None."""
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def _fmt_csv(v):
    """Format a cell value for CSV output."""
    if v is None:
        return ""
    return str(v)


def build_row(emp, contrib):
    """Build a 19-element data row from employee ref + (optional) contribution dict."""
    c = contrib or {"hours": 0, "earnings": 0, "pretax": 0, "roth": 0, "employer": 0}
    return [
        emp["ssn"],
        emp["first"],
        emp["mi"]    or None,
        emp["last"],
        emp["sex"]   or None,
        emp["dob"]   or None,
        emp["hire"]  or None,
        None,                        # termination date
        None,                        # rehire date
        _clean(c["earnings"]),
        _clean(c["hours"]),
        _clean(c["pretax"]),
        _clean(c["roth"]),
        _clean(c["employer"]),
        emp["addr1"] or None,
        emp["addr2"] or None,
        emp["city"]  or None,
        emp["state"] or None,
        emp["zip"]   or None,
    ]


def write_outputs(rows, check_date, output_dir):
    date_str = check_date.strftime("%Y%m%d")
    stem = (
        f"{META_PREFIX}-{date_str}"
        f"-001-CREATIVEPLANNING-DIRECTLIGHTING-GUID-{CP_GUID}"
    )

    # CSV — no header, no meta row
    csv_path = output_dir / f"{stem}.csv"
    with open(csv_path, "w") as f:
        for row in rows:
            f.write(",".join(_fmt_csv(v) for v in row) + "\n")

    # XLSX — meta ID in A1, then data rows
    xlsx_path = output_dir / f"{stem}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws.cell(row=1, column=1, value=stem)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            if val is not None:
                ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(xlsx_path)

    return csv_path, xlsx_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]

    # --build-ref: rebuild employee roster from a Creative Planning template
    if args and args[0] == "--build-ref":
        if len(args) < 2:
            print("Usage: --build-ref /path/to/template.xlsx")
            sys.exit(1)
        ref = build_ref_from_xlsx(args[1])
        save_ref(ref)
        print(f"Saved {len(ref['employees'])} employees to {REF_FILE}")
        for e in ref["employees"]:
            print(f"  {e['last4_ssn']}: {e['last']}, {e['first']}")
        return

    # Determine payroll folder
    if args:
        payroll_folder = Path(args[0])
    else:
        payroll_folder = find_latest_payroll_folder()
        if not payroll_folder:
            print("ERROR: Could not auto-detect payroll folder. Pass path as argument.")
            sys.exit(1)
        print(f"Payroll folder: {payroll_folder}")

    if not payroll_folder.exists():
        print(f"ERROR: Folder not found: {payroll_folder}")
        sys.exit(1)

    # Find Retirement Plan Summary PDF
    pdf_path = find_retirement_pdf(payroll_folder)
    if not pdf_path:
        print(f"ERROR: No Retirement Plan Summary PDF in {payroll_folder}")
        sys.exit(1)
    print(f"PDF: {pdf_path.name}")

    # Parse PDF
    check_date, contribs = parse_retirement_pdf(pdf_path)
    if not check_date:
        print("ERROR: Could not extract check date from PDF.")
        sys.exit(1)

    print(f"Check date: {check_date}")
    print(f"Participants in PDF: {len(contribs)}")

    # Load employee roster
    ref = load_ref()
    if not ref.get("employees"):
        print(f"ERROR: No employee reference at {REF_FILE}")
        print("Run: python3 paychex_to_401k.py --build-ref /path/to/template.xlsx")
        sys.exit(1)

    # Build rows — roster order, zeros for non-contributors
    rows = []
    print("\nEmployee contributions this period:")
    for emp in ref["employees"]:
        last4 = emp["last4_ssn"]
        c = contribs.get(last4)
        if c:
            print(f"  {emp['last']:15s} roth={c['roth']:7.2f}  pretax={c['pretax']:7.2f}  "
                  f"employer={c['employer']:6.2f}  earnings={c['earnings']:8.2f}")
        else:
            print(f"  {emp['last']:15s} (no contribution this period)")
        rows.append(build_row(emp, c))

    # Write CSV + XLSX
    csv_path, xlsx_path = write_outputs(rows, check_date, payroll_folder)
    print(f"\nGenerated:")
    print(f"  {csv_path}")
    print(f"  {xlsx_path}")


if __name__ == "__main__":
    main()
